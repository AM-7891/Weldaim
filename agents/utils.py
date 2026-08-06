"""
WELDAIM - utils.py
Funzioni di utilita condivise tra tutti gli agenti WeldAIM.

Importare con:
    from utils import (estrai_testo_pdf, estrai_testo_excel,
                       analizza_pdf_chunked, analizza_testo_chunked,
                       trova_file_per_estensione, pulisci_json)

Percorso: C:/Users/angma/Desktop/weldaim/agents/utils.py

Aggiornamento 2026-07-19:
- Aggiunto temperature=0 su tutte le chiamate client.messages.create() di questo file.
  Osservata variabilita' di estrazione tra run identici sullo stesso documento (es. campo
  disegno VT letto vuoto in un run e con contenuto in un run successivo, stesso PDF).
  temperature=0 riduce fortemente il campionamento casuale del modello, ma non lo azzera
  in modo garantito (Anthropic non garantisce determinismo assoluto nemmeno a temp=0).
  NOTA: questo fix NON risolve un problema distinto e ancora aperto — il chunking taglia
  per conteggio di caratteri/pagine senza consapevolezza della struttura del documento,
  il che puo' troncare righe di tabella a cavallo tra due chunk (visto su welding map,
  WM-01). Quel problema resta in roadmap post-demo (task #16), richiede un intervento
  piu' strutturale sulla logica di split, non risolvibile con questo parametro.

Aggiornamento 2026-07-28:
- Aggiunto prompt caching su tutte le chiamate client.messages.create() di questo file.
  PROBLEMA RISOLTO: prima di questo fix, ogni chiamata mandava prompt_per_chunk gia'
  fuso col testo del chunk in un'unica stringa (prompt_per_chunk.replace(...)). Il
  caching richiede blocchi identici byte-per-byte tra chiamate: fondendo fisso e
  variabile in una stringa unica, nessun blocco poteva mai risultare cacheabile.
  SOLUZIONE: la nuova funzione interna _prepara_content_con_cache() divide il
  prompt in due blocchi separati nella chiamata API - la parte fissa del template
  (istruzioni, criteri, schema JSON) marcata con cache_control, e il contenuto
  variabile (chunk di testo, o risultati parziali per l'aggregazione) non cacheato.
  Nessuna modifica richiesta ai file agente: continuano a passare prompt_per_chunk
  con il placeholder {testo_chunk} esattamente come prima.
  VINCOLO NOTO: il blocco cacheato deve superare 1024 token (~4000 caratteri) per
  Sonnet, altrimenti cache_control viene ignorato silenziosamente (nessun errore,
  nessun risparmio). Verificare hit reale dai contatori di utilizzo stampati ad
  ogni chiamata (cache_creation_input_tokens / cache_read_input_tokens), non
  assumere che il caching stia funzionando solo perche' il codice non da' errori.
  Cache TTL: 5 minuti di default, si rinnova ad ogni hit.
"""

import os
import re
import json
import fitz
import pytesseract
from pdf2image import convert_from_path
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# CONFIGURAZIONE GLOBALE
# ---------------------------------------------------------------------------

POPPLER_PATH = r"C:\Users\angma\Desktop\weldaim\poppler\Library\bin"
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

SOGLIA_TESTO_NATIVO = 50

PAGINE_PER_CHUNK_OCR = 4

CARATTERI_PER_CHUNK = 12000

# Temperatura fissata a 0 su tutte le chiamate al modello in questo file: riduce la
# variabilita' di estrazione tra run identici sullo stesso documento (non la elimina
# in modo garantito, ma la riduce fortemente rispetto al default).
TEMPERATURA_ESTRAZIONE = 0


# ---------------------------------------------------------------------------
# UTILITY 1 - pulizia JSON restituito da Claude
# ---------------------------------------------------------------------------

def pulisci_json(testo):
    """
    Rimuove eventuali backtick markdown dal JSON restituito da Claude.
    Usare sempre prima di json.loads() su risposte Claude.
    """
    if "```" in testo:
        match = re.search(r"```(?:json)?\s*(.*?)```", testo, re.DOTALL)
        if match:
            return match.group(1).strip()
        testo = testo.replace("```json", "").replace("```", "")
    return testo.strip()


# ---------------------------------------------------------------------------
# UTILITY INTERNA - costruzione content con prompt caching
# ---------------------------------------------------------------------------

def _prepara_content_con_cache(prompt_template, valore_dinamico, placeholder):
    """
    Divide un prompt template in blocchi separati per abilitare il prompt caching:
    - la parte fissa PRIMA del placeholder -> marcata cache_control (riusabile tra chiamate)
    - il valore dinamico + l'eventuale parte fissa DOPO il placeholder -> non cacheata

    Se il placeholder non e' presente nel template, ritorna un blocco unico senza
    cache_control (fallback sicuro, nessun comportamento diverso da prima).

    Nota: il blocco cacheato deve superare la soglia minima del modello (1024 token
    per Sonnet) per avere effetto reale. Sotto soglia, cache_control viene ignorato
    silenziosamente da Anthropic - nessun errore, nessun risparmio.
    """
    if placeholder not in prompt_template:
        return [{"type": "text", "text": prompt_template}]

    parte_fissa, parte_dopo = prompt_template.split(placeholder, 1)

    blocchi = []
    if parte_fissa.strip():
        blocchi.append({
            "type": "text",
            "text": parte_fissa,
            "cache_control": {"type": "ephemeral"}
        })
    blocchi.append({
        "type": "text",
        "text": valore_dinamico + parte_dopo
    })
    return blocchi


def _prepara_content_multi_cache(blocchi_cacheabili: list, blocco_finale: str):
    """
    Come _prepara_content_con_cache, ma supporta PIÙ breakpoint di cache
    nello stesso messaggio (Anthropic ne accetta fino a 4 per richiesta).

    Uso tipico: quando un prompt ha più parti statiche che cambiano a
    velocità diverse (es. istruzioni fisse = mai cambiano vs WPS/WPQR
    del welding book = statiche solo entro la stessa run).

    Parametri:
    - blocchi_cacheabili : lista di stringhe, ognuna diventa un blocco
                            con cache_control (marcata riusabile)
    - blocco_finale      : stringa dinamica, SEMPRE l'ultimo blocco,
                            mai cacheata (cambia ad ogni chiamata)

    Nota: ogni blocco cacheato sotto la soglia minima (1024 token per
    Sonnet) viene ignorato silenziosamente da Anthropic - nessun errore,
    nessun risparmio, ma nessun danno.
    """
    content = []
    for blocco in blocchi_cacheabili:
        if blocco.strip():
            content.append({
                "type": "text",
                "text": blocco,
                "cache_control": {"type": "ephemeral"}
            })
    content.append({"type": "text", "text": blocco_finale})
    return content

def _stampa_uso_cache(risposta, etichetta=""):
    """
    Stampa SEMPRE i contatori di utilizzo cache della risposta API (anche a zero),
    per distinguere "file non aggiornato / caching non attivo sul codice" da
    "codice attivo ma sotto soglia minima cacheabile per il modello".
    cache_creation_input_tokens: token scritti in cache in questa chiamata (costo 1.25x)
    cache_read_input_tokens: token letti dalla cache in questa chiamata (costo 0.1x)
    input_tokens: token input NON cacheati pagati a prezzo pieno in questa chiamata
    """
    try:
        uso = risposta.usage
        normali = getattr(uso, "input_tokens", 0) or 0
        creati = getattr(uso, "cache_creation_input_tokens", 0) or 0
        letti = getattr(uso, "cache_read_input_tokens", 0) or 0
        print(f"[CACHE{' ' + etichetta if etichetta else ''}] "
              f"input pieno: {normali} tok | scritti: {creati} tok | "
              f"letti dalla cache: {letti} tok")
    except Exception as e:
        print(f"[CACHE{' ' + etichetta if etichetta else ''}] "
              f"impossibile leggere usage: {e}")


# ---------------------------------------------------------------------------
# UTILITY 2 - estrazione testo da PDF completo
# ---------------------------------------------------------------------------

def estrai_testo_pdf(percorso_pdf):
    """
    Estrae testo da un PDF completo.
    Ritorna: (testo, is_ocr)
    - is_ocr=False: testo nativo PyMuPDF
    - is_ocr=True: testo OCR (PDF scansionato)
    Per PDF lunghi preferire analizza_pdf_chunked().
    """
    try:
        doc = fitz.open(percorso_pdf)
        testo = ""
        for pagina in doc:
            testo += pagina.get_text()
        doc.close()

        if len(testo.strip()) >= SOGLIA_TESTO_NATIVO:
            return testo.strip(), False

        print(f"[OCR] PDF scansionato: {os.path.basename(percorso_pdf)}")
        immagini = convert_from_path(percorso_pdf, dpi=300, poppler_path=POPPLER_PATH)
        testo_ocr = ""
        for img in immagini:
            testo_ocr += pytesseract.image_to_string(img, lang="ita+eng")
        return testo_ocr.strip(), True

    except Exception as e:
        return f"[ERRORE LETTURA PDF] {str(e)}", False


def estrai_testo_pdf_semplice(percorso_pdf):
    """
    Versione semplificata - ritorna solo il testo.
    Compatibile con Agent 1, 2, 3, 4 esistenti.
    """
    testo, _ = estrai_testo_pdf(percorso_pdf)
    return testo


# ---------------------------------------------------------------------------
# UTILITY 3 - estrazione testo da Excel
# ---------------------------------------------------------------------------

def estrai_testo_excel(percorso_excel):
    """
    Estrae contenuto da file Excel come testo leggibile da Claude.
    Legge tutti i fogli e tutte le celle non vuote.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(percorso_excel, data_only=True)
        righe_totali = []
        for nome_foglio in wb.sheetnames:
            ws = wb[nome_foglio]
            righe_totali.append(f"=== FOGLIO: {nome_foglio} ===")
            for riga in ws.iter_rows(values_only=True):
                valori = [str(v) if v is not None else "" for v in riga]
                if any(v.strip() for v in valori):
                    righe_totali.append("\t".join(valori))
        return "\n".join(righe_totali)
    except Exception as e:
        return f"[ERRORE LETTURA EXCEL] {str(e)}"


# ---------------------------------------------------------------------------
# UTILITY 4 - trova file per estensione in una cartella
# ---------------------------------------------------------------------------

def trova_file_per_estensione(cartella, estensioni):
    """
    Restituisce lista percorsi file nella cartella con le estensioni indicate.
    Ricerca case-insensitive. Non ricorsiva.
    """
    risultati = []
    if not os.path.isdir(cartella):
        return risultati
    for f in os.listdir(cartella):
        if any(f.lower().endswith(ext.lower()) for ext in estensioni):
            risultati.append(os.path.join(cartella, f))
    return risultati


# ---------------------------------------------------------------------------
# UTILITY 5 - chunking su testo gia estratto (PDF nativo, Excel, TXT)
# ---------------------------------------------------------------------------

def analizza_testo_chunked(
    testo,
    client,
    model,
    prompt_per_chunk,
    prompt_aggregazione,
    caratteri_per_chunk=CARATTERI_PER_CHUNK,
    max_tokens_chunk=1500,
    max_tokens_aggregazione=2000,
    nome_file="documento"
):
    """
    Analizza un testo lungo dividendolo in chunk da N caratteri.
    Usare per: PDF nativi lunghi, Excel lunghi, file TXT lunghi.
    NON usare per PDF scansionati (OCR) - usare analizza_pdf_chunked().
    """

    if len(testo) <= caratteri_per_chunk:
        print(f"[CHUNK-TESTO] {nome_file}: testo corto ({len(testo)} car.) - analisi diretta")
        content = _prepara_content_con_cache(prompt_per_chunk, testo, "{testo_chunk}")
        risposta = client.messages.create(
            model=model,
            max_tokens=max_tokens_chunk,
            temperature=TEMPERATURA_ESTRAZIONE,
            messages=[{"role": "user", "content": content}]
        )
        _stampa_uso_cache(risposta, nome_file)
        testo_risposta = pulisci_json(risposta.content[0].text)
        try:
            risultato = json.loads(testo_risposta)
            risultato["_chunking"] = {"usato": False, "caratteri": len(testo)}
            return risultato
        except json.JSONDecodeError:
            return {
                "raw": testo_risposta,
                "_chunking": {"usato": False},
                "nc": [{"severita": "ATTENZIONE", "codice": "TCHUNK-01",
                        "descrizione": f"Risposta non parsabile per {nome_file}",
                        "riferimento": "Errore interno"}]
            }

    # Divide il testo in chunk tagliando su interruzione di riga
    chunks = []
    inizio = 0
    while inizio < len(testo):
        fine = inizio + caratteri_per_chunk
        if fine >= len(testo):
            chunks.append(testo[inizio:])
            break
        taglio = testo.rfind("\n", inizio, fine)
        if taglio == -1 or taglio <= inizio:
            taglio = fine
        chunks.append(testo[inizio:taglio])
        inizio = taglio + 1

    print(f"[CHUNK-TESTO] {nome_file}: {len(testo)} car. -> {len(chunks)} chunk")

    risultati_parziali = []
    for i, chunk in enumerate(chunks):
        print(f"[CHUNK-TESTO] Analisi chunk {i + 1}/{len(chunks)}...")
        content = _prepara_content_con_cache(prompt_per_chunk, chunk, "{testo_chunk}")
        try:
            risposta = client.messages.create(
                model=model,
                max_tokens=max_tokens_chunk,
                temperature=TEMPERATURA_ESTRAZIONE,
                messages=[{"role": "user", "content": content}]
            )
            _stampa_uso_cache(risposta, f"{nome_file} chunk {i+1}")
            testo_risposta = pulisci_json(risposta.content[0].text)
            try:
                risultato_chunk = json.loads(testo_risposta)
            except json.JSONDecodeError:
                risultato_chunk = {
                    "raw": testo_risposta,
                    "nc": [{"severita": "APPUNTO", "codice": f"TCHUNK-{i+1:02d}",
                             "descrizione": f"Chunk {i+1} non parsabile",
                             "riferimento": "Errore interno"}]
                }
            risultati_parziali.append({"chunk": i + 1, "risultato": risultato_chunk})
        except Exception as e:
            risultati_parziali.append({"chunk": i + 1, "errore": str(e)})

    return _aggrega_risultati(
        risultati_parziali, client, model,
        prompt_aggregazione, max_tokens_aggregazione,
        meta={"usato": True, "caratteri": len(testo), "num_chunk": len(chunks)}
    )


# ---------------------------------------------------------------------------
# UTILITY 6 - chunking per pagina su PDF (nativo o scansionato)
# ---------------------------------------------------------------------------

def analizza_pdf_chunked(
    percorso_pdf,
    client,
    model,
    prompt_per_chunk,
    prompt_aggregazione,
    pagine_per_chunk=PAGINE_PER_CHUNK_OCR,
    max_tokens_chunk=1500,
    max_tokens_aggregazione=2000
):
    """
    Analizza un PDF lungo con strategia adattiva:
    - PDF nativo: estrae testo e usa analizza_testo_chunked()
    - PDF scansionato (OCR): chunking per pagina con OCR
    """
    nome_file = os.path.basename(percorso_pdf)

    try:
        doc = fitz.open(percorso_pdf)
        totale_pagine = len(doc)
    except Exception as e:
        return {
            "errore": str(e),
            "nc": [{"severita": "ATTENZIONE", "codice": "PCHUNK-00",
                    "descrizione": f"File {nome_file} non apribile: {str(e)}",
                    "riferimento": "Errore interno"}]
        }

    # Verifica se PDF e nativo o scansionato (campiona prime 3 pagine)
    testo_campione = ""
    for i in range(min(3, totale_pagine)):
        testo_campione += doc[i].get_text()

    is_ocr = len(testo_campione.strip()) < SOGLIA_TESTO_NATIVO * 3

    if not is_ocr:
        # PDF NATIVO - estrae tutto il testo e usa chunking su caratteri
        print(f"[CHUNK-PDF] {nome_file}: PDF nativo ({totale_pagine} pag.) -> chunking su testo")
        testo_completo = ""
        for pagina in doc:
            testo_completo += pagina.get_text()
        doc.close()
        return analizza_testo_chunked(
            testo=testo_completo,
            client=client,
            model=model,
            prompt_per_chunk=prompt_per_chunk,
            prompt_aggregazione=prompt_aggregazione,
            max_tokens_chunk=max_tokens_chunk,
            max_tokens_aggregazione=max_tokens_aggregazione,
            nome_file=nome_file
        )

    # PDF SCANSIONATO (OCR) - chunking per pagina
    print(f"[CHUNK-PDF] {nome_file}: PDF scansionato ({totale_pagine} pag.) -> "
          f"OCR chunk da {pagine_per_chunk} pagine")

    risultati_parziali = []
    num_chunk = 0

    for inizio in range(0, totale_pagine, pagine_per_chunk):
        fine = min(inizio + pagine_per_chunk, totale_pagine)
        num_chunk += 1
        print(f"[CHUNK-PDF] OCR chunk {num_chunk}: pagine {inizio+1}-{fine}")

        try:
            immagini = convert_from_path(
                percorso_pdf,
                dpi=300,
                poppler_path=POPPLER_PATH,
                first_page=inizio + 1,
                last_page=fine
            )
            testo_chunk = ""
            for j, img in enumerate(immagini):
                testo_pagina = pytesseract.image_to_string(img, lang="ita+eng")
                testo_chunk += f"\n--- Pagina {inizio + j + 1} ---\n{testo_pagina}"
        except Exception as e:
            risultati_parziali.append({
                "chunk": num_chunk,
                "pagine": f"{inizio+1}-{fine}",
                "errore": str(e)
            })
            continue

        content = _prepara_content_con_cache(
            prompt_per_chunk, testo_chunk.strip(), "{testo_chunk}"
        )
        try:
            risposta = client.messages.create(
                model=model,
                max_tokens=max_tokens_chunk,
                temperature=TEMPERATURA_ESTRAZIONE,
                messages=[{"role": "user", "content": content}]
            )
            _stampa_uso_cache(risposta, f"{nome_file} OCR chunk {num_chunk}")
            testo_risposta = pulisci_json(risposta.content[0].text)
            try:
                risultato_chunk = json.loads(testo_risposta)
            except json.JSONDecodeError:
                risultato_chunk = {
                    "raw": testo_risposta,
                    "nc": [{"severita": "APPUNTO", "codice": f"PCHUNK-{num_chunk:02d}",
                             "descrizione": f"Chunk OCR {num_chunk} non parsabile",
                             "riferimento": "Errore interno"}]
                }
            risultati_parziali.append({
                "chunk": num_chunk,
                "pagine": f"{inizio+1}-{fine}",
                "risultato": risultato_chunk
            })
        except Exception as e:
            risultati_parziali.append({
                "chunk": num_chunk,
                "pagine": f"{inizio+1}-{fine}",
                "errore": str(e)
            })

    doc.close()

    return _aggrega_risultati(
        risultati_parziali, client, model,
        prompt_aggregazione, max_tokens_aggregazione,
        meta={
            "usato": True,
            "tipo": "OCR",
            "pagine_totali": totale_pagine,
            "num_chunk": num_chunk,
            "pagine_per_chunk": pagine_per_chunk
        }
    )


# ---------------------------------------------------------------------------
# UTILITY INTERNA - aggregazione risultati parziali
# ---------------------------------------------------------------------------

def _aggrega_risultati(
    risultati_parziali,
    client,
    model,
    prompt_aggregazione,
    max_tokens,
    meta
):
    """
    Funzione interna - aggrega i risultati parziali dei chunk
    in un unico JSON finale tramite Claude.
    """
    print(f"[CHUNK] Aggregazione {len(risultati_parziali)} chunk...")

    valore_risultati = json.dumps(risultati_parziali, ensure_ascii=False, indent=2)
    content = _prepara_content_con_cache(
        prompt_aggregazione, valore_risultati, "{risultati_parziali}"
    )

    try:
        risposta = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            temperature=TEMPERATURA_ESTRAZIONE,
            messages=[{"role": "user", "content": content}]
        )
        _stampa_uso_cache(risposta, "aggregazione")
        testo = pulisci_json(risposta.content[0].text)
        try:
            risultato_finale = json.loads(testo)
        except json.JSONDecodeError:
            risultato_finale = {
                "raw": testo,
                "nc": [{"severita": "ATTENZIONE", "codice": "AGG-01",
                        "descrizione": "Aggregazione chunk non parsabile",
                        "riferimento": "Errore interno"}]
            }
    except Exception as e:
        risultato_finale = {
            "errore_aggregazione": str(e),
            "nc": [{"severita": "ATTENZIONE", "codice": "AGG-00",
                    "descrizione": f"Errore aggregazione: {str(e)}",
                    "riferimento": "Errore interno"}]
        }

    risultato_finale["_chunking"] = meta
    return risultato_finale